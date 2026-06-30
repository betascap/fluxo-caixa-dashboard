"""
Le o relatorio "Contas Pagas (por Centro de Custo)" exportado do Sienge (PDF)
e lanca os totais liquidos por centro de custo na aba de Fluxo de Caixa do
arquivo FCVILLE, na coluna do mes correspondente.

Uso:
    python sienge_to_fc.py <relatorio.pdf> <fcville.xlsx> "<nome da aba>"

Exemplo:
    python sienge_to_fc.py "CONTAS PAGAS.pdf" "FCVILLE MAIO 2026.xlsx" "FC Total (10 lotes)"
"""
import re
import sys
import datetime
import pdfplumber
import openpyxl


# Centro de custo (Sienge) -> linha do FC. Centros 5001 e 5003 sao somados
# na mesma linha (Comissoes S/ Venda), pois o template nao separa equipe
# interna de corretagem externa.
CC_TO_ROW = {
    "5001": 25,  # Receita de Vendas (corretagem externa) -> Comissoes S/ Venda
    "5003": 25,  # Comercial (equipe interna)              -> Comissoes S/ Venda
    "5002": 50,  # Administrativo                          -> Outras Despesas (adm)
    "5004": 27,  # Marketing                                -> Marketing
    "5005": 26,  # Stand de Vendas                          -> Stand de Venda - Manutencao
    "5006": 43,  # Incorporacao                             -> Incorporacao
    "5007": 38,  # Custos de Obra                           -> OBRA (revisao ABRIL/26)
    "5009": 33,  # Terreno                                  -> IPTU Estoque
    "5014": 56,  # Impostos                                 -> Impostos
    "5015": 32,  # Clientes (ITBI)                          -> ITBI
}

MESES_PT = {
    1: "janeiro", 2: "fevereiro", 3: "marco", 4: "abril", 5: "maio", 6: "junho",
    7: "julho", 8: "agosto", 9: "setembro", 10: "outubro", 11: "novembro", 12: "dezembro",
}

CC_HEADER_RE = re.compile(r"Centro de custo\s*(\d{4})\s*-")
TOTAL_CC_RE = re.compile(
    r"Total do centro de custo\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)"
)
PERIODO_RE = re.compile(r"Per.odo\s*(\d{2}/\d{2}/\d{4})\s+a\s+(\d{2}/\d{2}/\d{4})")


def br_to_float(s: str) -> float:
    return float(s.replace(".", "").replace(",", "."))


def parse_sienge_pdf(pdf_path: str):
    """Retorna (periodo_inicio: date, totais_liquidos: {cc: float})."""
    totais = {}
    periodo_inicio = None

    with pdfplumber.open(pdf_path) as pdf:
        full_text = "\n".join(page.extract_text() or "" for page in pdf.pages)

    if periodo_inicio is None:
        m = PERIODO_RE.search(full_text)
        if m:
            periodo_inicio = datetime.datetime.strptime(m.group(1), "%d/%m/%Y").date()

    cc_atual = None
    for line in full_text.splitlines():
        cc_match = CC_HEADER_RE.search(line)
        if cc_match:
            cc_atual = cc_match.group(1)
            continue

        total_match = TOTAL_CC_RE.search(line)
        if total_match and cc_atual:
            liquido = br_to_float(total_match.group(4))
            totais[cc_atual] = totais.get(cc_atual, 0.0) + liquido
            cc_atual = None  # evita recontagem ate o proximo cabecalho

    return periodo_inicio, totais


def find_month_column(ws, ano: int, mes: int, header_row: int = 5, max_col: int = 320):
    for c in range(1, max_col + 1):
        v = ws.cell(row=header_row, column=c).value
        if isinstance(v, datetime.datetime) and v.year == ano and v.month == mes:
            return c
    return None


def main():
    if len(sys.argv) != 4:
        print(__doc__)
        sys.exit(1)

    pdf_path, xlsx_path, sheet_name = sys.argv[1], sys.argv[2], sys.argv[3]

    periodo_inicio, totais = parse_sienge_pdf(pdf_path)
    if periodo_inicio is None:
        print("ERRO: nao consegui identificar o periodo do relatorio no PDF.")
        sys.exit(1)

    print(f"Periodo identificado no relatorio: {MESES_PT[periodo_inicio.month]}/{periodo_inicio.year}")
    print("Totais liquidos por centro de custo:")
    for cc, val in sorted(totais.items()):
        print(f"  {cc}: R$ {val:,.2f}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    if sheet_name not in wb.sheetnames:
        print(f"ERRO: aba '{sheet_name}' nao encontrada. Abas disponiveis: {wb.sheetnames}")
        sys.exit(1)
    ws = wb[sheet_name]

    col = find_month_column(ws, periodo_inicio.year, periodo_inicio.month)
    if col is None:
        print(f"ERRO: nao encontrei a coluna do mes {periodo_inicio.month}/{periodo_inicio.year} na aba '{sheet_name}'.")
        sys.exit(1)
    col_letter = openpyxl.utils.get_column_letter(col)
    print(f"Coluna do mes no FC: {col_letter}")

    # soma por linha de destino (caso dois CCs apontem pra mesma linha, ex. 5001+5003)
    valores_por_linha = {}
    for cc, valor in totais.items():
        row = CC_TO_ROW.get(cc)
        if row is None:
            continue  # centro de custo sem mapeamento (ex.: 5015 Clientes ja mapeado, mas outros podem nao estar)
        valores_por_linha[row] = valores_por_linha.get(row, 0.0) + valor

    print("\nLancamentos no FC:")
    for row, valor in sorted(valores_por_linha.items()):
        label = ws.cell(row=row, column=1).value
        valor_negativo = -abs(valor)  # despesas sao sempre negativas no template
        ws.cell(row=row, column=col, value=valor_negativo)
        print(f"  Linha {row} ({label}) -> {col_letter}{row} = {valor_negativo:,.2f}")

    wb.save(xlsx_path)
    print(f"\nArquivo salvo: {xlsx_path}")


if __name__ == "__main__":
    main()
