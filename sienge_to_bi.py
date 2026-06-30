"""
Le o relatorio "Contas Pagas (por Centro de Custo)" (PDF) e gera:
  1. CSV estruturado para Power BI (long format)
  2. Excel atualizado com FC (compatibilidade)

Uso:
    python sienge_to_bi.py <relatorio.pdf> [fcville.xlsx] [output.csv]

Exemplo:
    python sienge_to_bi.py "CONTAS PAGAS.pdf" "FCVILLE.xlsx" "fc_dados.csv"

Se nao passar Excel/CSV, usa defaults: FCVILLE_OUTPUT.xlsx e fc_dados.csv
"""
import re
import sys
import csv
import datetime
import pdfplumber
import openpyxl


CC_TO_ROW = {
    "5001": (25, "Receita de Vendas", "Comissões S/ Venda (Equipe Interna)"),
    "5003": (25, "Comercial", "Comissões S/ Venda (Equipe Interna)"),
    "5002": (50, "Administrativo", "Outras Despesas (adm)"),
    "5004": (27, "Marketing", "Marketing"),
    "5005": (26, "Stand de Vendas", "Stand de Venda - Manutenção"),
    "5006": (43, "Incorporação", "Incorporação"),
    "5007": (38, "Custos de Obra", "OBRA - (revisão ABRIL/26)"),
    "5009": (33, "Terreno", "IPTU Estoque"),
    "5014": (56, "Impostos", "Impostos"),
    "5015": (32, "Clientes (ITBI)", "ITBI"),
}

MESES_PT = {
    1: "janeiro", 2: "fevereiro", 3: "marco", 4: "abril", 5: "maio", 6: "junho",
    7: "julho", 8: "agosto", 9: "setembro", 10: "outubro", 11: "novembro", 12: "dezembro",
}

CC_HEADER_RE = re.compile(r"Centro de custo\s*(\d{4})\s*-")
TOTAL_CC_RE = re.compile(
    r"Total do centro de custo\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)"
)
PERIODO_RE = re.compile(r"Per.odo\s*(\d{2}/\d{2}/\d{4})\s+a\s+(\d{2}/\d{2}/\d{4})")


def br_to_float(s: str) -> float:
    return float(s.replace(".", "").replace(",", "."))


def parse_sienge_pdf(pdf_path: str):
    totais = {}
    periodo_inicio = None

    with pdfplumber.open(pdf_path) as pdf:
        full_text = "\n".join(page.extract_text() or "" for page in pdf.pages)

    if periodo_inicio is None:
        m = PERIODO_RE.search(full_text)
        if m:
            periodo_inicio = datetime.datetime.strptime(m.group(1), "%d/%m/%Y").date()

    cc_atual = None
    for line in full_text.splitlines():
        cc_match = CC_HEADER_RE.search(line)
        if cc_match:
            cc_atual = cc_match.group(1)
            continue

        total_match = TOTAL_CC_RE.search(line)
        if total_match and cc_atual:
            liquido = br_to_float(total_match.group(4))
            totais[cc_atual] = totais.get(cc_atual, 0.0) + liquido
            cc_atual = None

    return periodo_inicio, totais


def find_month_column(ws, ano: int, mes: int, header_row: int = 5, max_col: int = 320):
    for c in range(1, max_col + 1):
        v = ws.cell(row=header_row, column=c).value
        if isinstance(v, datetime.datetime) and v.year == ano and v.month == mes:
            return c
    return None


def gerar_csv(periodo_inicio: datetime.date, totais: dict, csv_path: str):
    """Gera CSV em formato long para Power BI."""
    rows = []
    for cc in sorted(totais.keys()):
        if cc not in CC_TO_ROW:
            continue
        row_num, categoria, descricao = CC_TO_ROW[cc]
        valor = totais[cc]
        rows.append({
            "Data": periodo_inicio.strftime("%Y-%m-%d"),
            "CentroCusto": cc,
            "Categoria": categoria,
            "DescricaoLinha": descricao,
            "Valor": valor,
        })

    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["Data", "CentroCusto", "Categoria", "DescricaoLinha", "Valor"])
        writer.writeheader()
        writer.writerows(rows)

    return csv_path


def atualizar_excel(periodo_inicio: datetime.date, totais: dict, xlsx_path: str, sheet_name: str = "FC Total (10 lotes)"):
    """Atualiza Excel com compatibilidade (opcional)."""
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=False)
        if sheet_name not in wb.sheetnames:
            print(f"AVISO: aba '{sheet_name}' nao encontrada em {xlsx_path}. Pulando atualizacao Excel.")
            return
        ws = wb[sheet_name]

        col = find_month_column(ws, periodo_inicio.year, periodo_inicio.month)
        if col is None:
            print(f"AVISO: coluna do mes {periodo_inicio.month}/{periodo_inicio.year} nao encontrada. Pulando.")
            return

        valores_por_linha = {}
        for cc, valor in totais.items():
            if cc not in CC_TO_ROW:
                continue
            row = CC_TO_ROW[cc][0]
            valores_por_linha[row] = valores_por_linha.get(row, 0.0) + valor

        for row, valor in valores_por_linha.items():
            valor_negativo = -abs(valor)
            ws.cell(row=row, column=col, value=valor_negativo)

        wb.save(xlsx_path)
        print(f"Excel atualizado: {xlsx_path}")
    except Exception as e:
        print(f"AVISO: erro ao atualizar Excel: {e}")


def main():
    pdf_path = sys.argv[1] if len(sys.argv) > 1 else "CONTAS PAGAS.pdf"
    xlsx_path = sys.argv[2] if len(sys.argv) > 2 else "FCVILLE_OUTPUT.xlsx"
    csv_path = sys.argv[3] if len(sys.argv) > 3 else "fc_dados.csv"

    if not __file__.endswith("sienge_to_bi.py"):
        print(__doc__)

    periodo_inicio, totais = parse_sienge_pdf(pdf_path)
    if periodo_inicio is None:
        print("ERRO: nao consegui identificar o periodo do relatorio no PDF.")
        sys.exit(1)

    print(f"Periodo: {MESES_PT[periodo_inicio.month]}/{periodo_inicio.year}")
    print("Totais liquidos por centro de custo:")
    for cc in sorted(totais.keys()):
        val = totais[cc]
        print(f"  {cc}: R$ {val:,.2f}")

    print(f"\nGerando CSV: {csv_path}")
    gerar_csv(periodo_inicio, totais, csv_path)
    print(f"CSV criado com sucesso.")

    print(f"\nAtualizando Excel (compatibilidade): {xlsx_path}")
    atualizar_excel(periodo_inicio, totais, xlsx_path)

    print("\nDone! Proximos passos:")
    print(f"  1. Importe {csv_path} no Power BI Desktop")
    print(f"  2. Recarregue o template Power BI")
    print(f"  3. Publique no Power BI Service")


if __name__ == "__main__":
    main()
